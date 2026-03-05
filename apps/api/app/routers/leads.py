"""Leads Router — Slice 95.

9 endpoints for the Lead Gen CRM (Clay/Apollo-style):
  GET    /leads                  — list leads (brand_id, status filter)
  POST   /leads                  — create single lead manually
  POST   /leads/generate         — AI-generate N leads from ICP (cap 20)
  POST   /leads/batch-enrich     — enrich pasted list (cap 3 — Vercel timeout)
  POST   /leads/enrich/{id}      — full 3-engine enrich + BANT score
  POST   /leads/outreach/{id}    — generate icebreaker + DM + email + sequence
  PATCH  /leads/{id}             — update status/notes/transcript/icebreaker/sequence
  DELETE /leads/{id}             — remove lead
  GET    /leads/export           — download as .xlsx (Instantly.ai-compatible)

Security: A01 IDOR (.eq user_id), A03 UUID validation, A07 JWT, A10 SSRF, A05 private data.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from fastapi.responses import Response
from pydantic import BaseModel, Field

from app.auth import CurrentUser, get_current_user
from app.deps import get_admin_client
from app.services.lead_gen import enrich_lead, generate_leads_from_icp, generate_outreach, research_icp, ICP_METHODOLOGY

logger = logging.getLogger("app.routers.leads")

router = APIRouter(tags=["leads"])

_UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
    re.IGNORECASE,
)

# Sanitise name inputs — allow word chars, spaces, hyphens, dots, commas
_SAFE_NAME_RE = re.compile(r"[^\w\s\-\.\,]")


# ── Schemas ────────────────────────────────────────────────────────────────


class LeadCreate(BaseModel):
    brand_id: str
    full_name: str = Field(..., min_length=1, max_length=255)
    title: Optional[str] = Field(default=None, max_length=255)
    company: Optional[str] = Field(default=None, max_length=255)
    email: Optional[str] = Field(default=None, max_length=255)
    linkedin_url: Optional[str] = Field(default=None, max_length=500)
    company_website: Optional[str] = Field(default=None, max_length=500)
    location: Optional[str] = Field(default=None, max_length=255)
    notes: Optional[str] = None
    source: str = "manual"


class LeadUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None
    transcript: Optional[str] = None
    icebreaker: Optional[str] = None
    sequence: Optional[List[Dict[str, Any]]] = None


class GenerateRequest(BaseModel):
    brand_id: str
    count: int = Field(default=10, ge=1, le=20)


class BatchEnrichItem(BaseModel):
    full_name: str
    company: Optional[str] = None
    linkedin_url: Optional[str] = None
    email: Optional[str] = None


class BatchEnrichRequest(BaseModel):
    brand_id: str
    leads: List[BatchEnrichItem] = Field(..., min_length=1, max_length=3)


# ── Helpers ────────────────────────────────────────────────────────────────


def _row_to_dict(row: Dict) -> Dict:
    """Normalise a DB row for API response."""
    return {
        "id": row["id"],
        "user_id": row["user_id"],
        "brand_id": row["brand_id"],
        "full_name": row.get("full_name", ""),
        "title": row.get("title"),
        "company": row.get("company"),
        "linkedin_url": row.get("linkedin_url"),
        "company_website": row.get("company_website"),
        "location": row.get("location"),
        "email": row.get("email"),
        "twitter_handle": row.get("twitter_handle"),
        "status": row.get("status", "cold"),
        "source": row.get("source", "manual"),
        "enrichment": row.get("enrichment") or {},
        "bant_score": row.get("bant_score") or 0,
        "notes": row.get("notes"),
        "transcript": row.get("transcript"),
        "icebreaker": row.get("icebreaker"),
        "outreach_draft": row.get("outreach_draft") or {},
        "sequence": row.get("sequence") or [],
        "last_enriched_at": row.get("last_enriched_at"),
        "created_at": str(row.get("created_at", "")),
        "updated_at": str(row.get("updated_at", "")),
    }


VALID_STATUSES = {"cold", "warm", "hot", "customer", "disqualified"}


# ── Endpoints ──────────────────────────────────────────────────────────────


@router.get("/leads")
async def list_leads(
    brand_id: str = Query(...),
    status: Optional[str] = Query(default=None),
    user: CurrentUser = Depends(get_current_user),
) -> List[Dict]:
    """List leads for a brand. Optional status filter."""
    if not _UUID_RE.match(brand_id):
        raise HTTPException(400, "Invalid brand_id")

    sb = get_admin_client()
    q = (
        sb.table("leads")
        .select("*")
        .eq("brand_id", brand_id)
        .eq("user_id", user.id)
    )
    if status:
        q = q.eq("status", status)

    result = q.order("bant_score", desc=True).order("created_at", desc=True).execute()
    return [_row_to_dict(r) for r in (result.data or [])]


@router.post("/leads", status_code=201)
async def create_lead(
    body: LeadCreate,
    user: CurrentUser = Depends(get_current_user),
) -> Dict:
    """Create a single lead manually."""
    if not _UUID_RE.match(body.brand_id):
        raise HTTPException(400, "Invalid brand_id")

    # Sanitise name
    safe_name = _SAFE_NAME_RE.sub("", body.full_name).strip()
    if not safe_name:
        raise HTTPException(400, "full_name contains invalid characters")

    sb = get_admin_client()
    row = {
        "user_id": user.id,
        "brand_id": body.brand_id,
        "full_name": safe_name[:255],
        "title": (body.title or "")[:255] or None,
        "company": (body.company or "")[:255] or None,
        "email": (body.email or "")[:255] or None,
        "linkedin_url": (body.linkedin_url or "")[:500] or None,
        "company_website": (body.company_website or "")[:500] or None,
        "location": (body.location or "")[:255] or None,
        "notes": body.notes,
        "source": "manual",
        "status": "cold",
    }

    try:
        result = sb.table("leads").insert(row).execute()
    except Exception as exc:
        # Unique constraint violation = duplicate lead
        if "unique" in str(exc).lower() or "duplicate" in str(exc).lower():
            raise HTTPException(409, "A lead with this name and company already exists")
        raise HTTPException(500, "Failed to create lead")

    if not result.data:
        raise HTTPException(500, "Failed to create lead")

    return _row_to_dict(result.data[0])


@router.post("/leads/generate")
async def generate_leads(
    body: GenerateRequest,
    user: CurrentUser = Depends(get_current_user),
) -> List[Dict]:
    """AI-generate N leads from the brand's ICP. Cap: 20."""
    if not _UUID_RE.match(body.brand_id):
        raise HTTPException(400, "Invalid brand_id")

    try:
        leads = generate_leads_from_icp(body.brand_id, user.id, body.count)
    except ValueError as exc:
        raise HTTPException(422, str(exc))
    except Exception as exc:
        logger.warning("generate_leads_from_icp failed: %s", exc)
        raise HTTPException(500, "Lead generation failed — check your ICP settings")

    if not leads:
        raise HTTPException(422, "No leads found — try refining your ICP in Brand Profile")

    # Upsert all leads (dedup by name+company)
    sb = get_admin_client()
    saved = []
    for lead in leads:
        row = {
            "user_id": user.id,
            "brand_id": body.brand_id,
            **lead,
        }
        try:
            result = (
                sb.table("leads")
                .upsert(row, on_conflict="user_id,brand_id,full_name,company")
                .execute()
            )
            if result.data:
                saved.append(_row_to_dict(result.data[0]))
        except Exception as exc:
            logger.debug("Lead upsert skipped (likely duplicate): %s", exc)

    return saved


@router.post("/leads/batch-enrich")
async def batch_enrich(
    body: BatchEnrichRequest,
    user: CurrentUser = Depends(get_current_user),
) -> List[Dict]:
    """Enrich a pasted list of leads (cap: 3 to stay within Vercel 60s timeout)."""
    if not _UUID_RE.match(body.brand_id):
        raise HTTPException(400, "Invalid brand_id")

    sb = get_admin_client()
    results = []

    for item in body.leads[:3]:
        safe_name = _SAFE_NAME_RE.sub("", item.full_name).strip()
        if not safe_name:
            continue

        # Create or get the lead
        lead_row = {
            "user_id": user.id,
            "brand_id": body.brand_id,
            "full_name": safe_name[:255],
            "company": (item.company or "")[:255] or None,
            "linkedin_url": (item.linkedin_url or "")[:500] or None,
            "email": (item.email or "")[:255] or None,
            "source": "imported",
            "status": "cold",
        }

        try:
            upsert_result = (
                sb.table("leads")
                .upsert(lead_row, on_conflict="user_id,brand_id,full_name,company")
                .execute()
            )
            if not upsert_result.data:
                continue
            saved = upsert_result.data[0]
        except Exception as exc:
            logger.debug("Batch enrich upsert failed: %s", exc)
            continue

        # Enrich
        try:
            enriched = enrich_lead(saved)
            updates = {
                "enrichment": enriched["enrichment"],
                "bant_score": enriched["bant_score"],
                "last_enriched_at": enriched["enrichment"].get("last_enriched_at"),
            }
            update_result = (
                sb.table("leads")
                .update(updates)
                .eq("id", saved["id"])
                .eq("user_id", user.id)
                .execute()
            )
            if update_result.data:
                results.append(_row_to_dict(update_result.data[0]))
            else:
                results.append(_row_to_dict(saved))
        except Exception as exc:
            logger.warning("Enrich failed for %s: %s", safe_name, exc)
            results.append(_row_to_dict(saved))

    return results


@router.post("/leads/enrich/{lead_id}")
async def enrich_lead_endpoint(
    lead_id: str,
    user: CurrentUser = Depends(get_current_user),
) -> Dict:
    """Full 3-engine enrichment + BANT scoring for a single lead."""
    if not _UUID_RE.match(lead_id):
        raise HTTPException(400, "Invalid lead_id")

    sb = get_admin_client()
    result = (
        sb.table("leads")
        .select("*")
        .eq("id", lead_id)
        .eq("user_id", user.id)
        .limit(1)
        .execute()
    )
    if not result.data:
        raise HTTPException(404, "Lead not found")

    lead = result.data[0]

    try:
        enriched = enrich_lead(lead)
    except Exception as exc:
        logger.warning("enrich_lead failed for %s: %s", lead_id, exc)
        raise HTTPException(500, "Enrichment failed — please try again")

    updates = {
        "enrichment": enriched["enrichment"],
        "bant_score": enriched["bant_score"],
        "last_enriched_at": enriched["enrichment"].get("last_enriched_at"),
    }
    update_result = (
        sb.table("leads")
        .update(updates)
        .eq("id", lead_id)
        .eq("user_id", user.id)
        .execute()
    )
    if not update_result.data:
        raise HTTPException(500, "Failed to save enrichment")

    return _row_to_dict(update_result.data[0])


@router.post("/leads/outreach/{lead_id}")
async def generate_outreach_endpoint(
    lead_id: str,
    user: CurrentUser = Depends(get_current_user),
) -> Dict:
    """Generate icebreaker + DM + cold email + 3-message sequence for a lead."""
    if not _UUID_RE.match(lead_id):
        raise HTTPException(400, "Invalid lead_id")

    sb = get_admin_client()

    # 1. Fetch lead (IDOR guard)
    lead_result = (
        sb.table("leads")
        .select("*")
        .eq("id", lead_id)
        .eq("user_id", user.id)
        .limit(1)
        .execute()
    )
    if not lead_result.data:
        raise HTTPException(404, "Lead not found")
    lead = lead_result.data[0]

    # 2. Fetch brand profile (IDOR guard)
    brand_result = (
        sb.table("personal_brands")
        .select("profile_json")
        .eq("id", lead["brand_id"])
        .eq("user_id", user.id)
        .limit(1)
        .execute()
    )
    brand_profile = {}
    if brand_result.data:
        brand_profile = brand_result.data[0].get("profile_json") or {}

    # 3. Generate
    try:
        outreach = generate_outreach(lead, lead.get("enrichment") or {}, brand_profile)
    except Exception as exc:
        logger.warning("generate_outreach failed for %s: %s", lead_id, exc)
        raise HTTPException(500, "Outreach generation failed — please try again")

    # 4. Save to DB
    updates = {
        "icebreaker": outreach["icebreaker"],
        "outreach_draft": outreach["outreach_draft"],
        "sequence": outreach["sequence"],
    }
    update_result = (
        sb.table("leads")
        .update(updates)
        .eq("id", lead_id)
        .eq("user_id", user.id)
        .execute()
    )
    if not update_result.data:
        raise HTTPException(500, "Failed to save outreach")

    return _row_to_dict(update_result.data[0])


@router.patch("/leads/{lead_id}")
async def update_lead(
    lead_id: str,
    body: LeadUpdate,
    user: CurrentUser = Depends(get_current_user),
) -> Dict:
    """Update lead status, notes, transcript, icebreaker, or sequence."""
    if not _UUID_RE.match(lead_id):
        raise HTTPException(400, "Invalid lead_id")

    updates: Dict[str, Any] = {}

    if body.status is not None:
        if body.status not in VALID_STATUSES:
            raise HTTPException(400, f"Invalid status. Must be one of: {', '.join(VALID_STATUSES)}")
        updates["status"] = body.status

    if body.notes is not None:
        updates["notes"] = body.notes

    if body.transcript is not None:
        updates["transcript"] = body.transcript

    if body.icebreaker is not None:
        updates["icebreaker"] = body.icebreaker

    if body.sequence is not None:
        # Validate sequence items have required fields
        validated_seq = []
        for item in body.sequence:
            if isinstance(item, dict) and "label" in item and "message" in item:
                validated_seq.append({
                    "label": str(item.get("label", "")),
                    "day": int(item.get("day", 1)),
                    "channel": str(item.get("channel", "linkedin")),
                    "message": str(item.get("message", "")),
                    "sent_at": item.get("sent_at"),
                })
        updates["sequence"] = validated_seq

    if not updates:
        raise HTTPException(400, "No fields to update")

    sb = get_admin_client()
    result = (
        sb.table("leads")
        .update(updates)
        .eq("id", lead_id)
        .eq("user_id", user.id)
        .execute()
    )
    if not result.data:
        raise HTTPException(404, "Lead not found")

    return _row_to_dict(result.data[0])


@router.delete("/leads/{lead_id}", response_class=Response)
async def delete_lead(
    lead_id: str,
    user: CurrentUser = Depends(get_current_user),
):
    """Remove a lead."""
    if not _UUID_RE.match(lead_id):
        raise HTTPException(400, "Invalid lead_id")

    sb = get_admin_client()
    result = (
        sb.table("leads")
        .delete()
        .eq("id", lead_id)
        .eq("user_id", user.id)
        .execute()
    )
    if not result.data:
        raise HTTPException(404, "Lead not found")

    return Response(status_code=204)


@router.get("/leads/export")
async def export_leads(
    brand_id: str = Query(...),
    user: CurrentUser = Depends(get_current_user),
) -> Response:
    """Export leads as .xlsx (Instantly.ai-compatible format).

    Columns: Name, Title, Company, LinkedIn, Email, Location, Status,
             BANT Score, Professional Topics, Recent Achievements,
             Hiring Signals, Pain Points, Icebreaker, Created At

    Transcripts and notes are NEVER exported (OWASP A05 — private data).
    """
    if not _UUID_RE.match(brand_id):
        raise HTTPException(400, "Invalid brand_id")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(500, "Excel export not available (openpyxl not installed)")

    sb = get_admin_client()
    result = (
        sb.table("leads")
        .select("*")
        .eq("brand_id", brand_id)
        .eq("user_id", user.id)
        .neq("status", "disqualified")
        .order("bant_score", desc=True)
        .execute()
    )

    rows = result.data or []

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "First Name", "Last Name", "Title", "Company", "LinkedIn URL",
        "Email", "Location", "Status", "BANT Score",
        "Professional Topics", "Recent Achievements",
        "Hiring Signals", "Pain Points",
        "Icebreaker",  # Instantly.ai custom variable
        "Created At",
    ]

    # Header row styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for lead in rows:
        enrichment = lead.get("enrichment") or {}
        full_name = lead.get("full_name", "")
        name_parts = full_name.split(" ", 1)
        first_name = name_parts[0] if name_parts else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""

        def _join(field: str) -> str:
            val = enrichment.get(field) or []
            if isinstance(val, list):
                return "; ".join(str(v) for v in val[:3])
            return str(val)

        ws.append([
            first_name,
            last_name,
            lead.get("title", ""),
            lead.get("company", ""),
            lead.get("linkedin_url", ""),
            lead.get("email", ""),
            lead.get("location", ""),
            lead.get("status", "cold"),
            lead.get("bant_score", 0),
            _join("professional_topics"),
            _join("recent_achievements"),
            _join("hiring_signals"),
            _join("pain_points"),
            lead.get("icebreaker", ""),
            str(lead.get("created_at", ""))[:10],
        ])

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=leads.xlsx",
            "Cache-Control": "no-store",
        },
    )


# ── ICP Research ────────────────────────────────────────────────────────────


class IcpResearchRequest(BaseModel):
    brand_id: str
    product_name: Optional[str] = None
    pricing: Optional[str] = None
    platform: Optional[str] = None
    lead_database: str = "Apollo.io"
    scraping_tool: str = "Apify"


@router.post("/leads/icp-research")
async def run_icp_research(
    body: IcpResearchRequest,
    user: CurrentUser = Depends(get_current_user),
) -> Dict:
    """4-stage ICP research using the Sales Lead Research System Prompt methodology.

    Stage 1: Objective  — derived from brand profile
    Stage 2: Brand Snapshot — built from profile_json fields
    Stage 3: Research Questions — Perplexity searches for target companies + ICP signals
    Stage 4: Apollo Filters — company + contact + keyword filters ready to paste
    """
    if not _UUID_RE.match(body.brand_id):
        raise HTTPException(400, "Invalid brand_id")

    # IDOR guard
    sb = get_admin_client()
    brand_check = (
        sb.table("personal_brands")
        .select("id")
        .eq("id", body.brand_id)
        .eq("user_id", user.id)
        .limit(1)
        .execute()
    )
    if not brand_check.data:
        raise HTTPException(404, "Brand not found")

    try:
        result = research_icp(
            brand_id=body.brand_id,
            user_id=user.id,
            overrides={
                "product_name": body.product_name,
                "pricing": body.pricing,
                "platform": body.platform,
                "lead_database": body.lead_database,
                "scraping_tool": body.scraping_tool,
            },
        )
        return result
    except ValueError as exc:
        raise HTTPException(404, str(exc))
    except Exception as exc:
        logger.error("ICP research failed: %s", exc)
        raise HTTPException(500, "ICP research failed — check logs")


@router.get("/leads/icp-methodology")
async def get_icp_methodology(
    user: CurrentUser = Depends(get_current_user),
) -> Dict:
    """Return the ICP methodology template as plain text for display and agent seeding."""
    return {"content": ICP_METHODOLOGY, "title": "Sales Lead Research System Prompt Template"}
